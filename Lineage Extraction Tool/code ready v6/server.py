"""
Lineage Extraction Tool — Redesigned FastAPI Server
=====================================================
Implements manager feedback:
  1. Column-level lineage (not table-level)
  2. file_path replaces domain as primary identifier
  3. Removed: UDF, delta, complexity, domain, subject_area, schedule, owner
  4. Sheet 1: Column Lineage  |  Sheet 2: AI Job Summary
  5. Two input modes: file upload  +  Git repository

Endpoints:
  POST /extract/upload        → upload files directly, returns .xlsx
  POST /extract/from-git      → connect to Git repo, fetch files, returns .xlsx
  GET  /health                → {"status": "ok"}
  GET  /                      → serves the HTML UI

Usage:
  pip install fastapi uvicorn python-multipart httpx gitpython openpyxl pydantic
  uvicorn server:app --host 0.0.0.0 --port 8000 --reload
"""

import io
import json
import os
import shutil
import sys
import tempfile
import traceback
from pathlib import Path
from typing import List, Optional

import httpx
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

# ---------------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))

# Add every candidate sub-directory that might hold the reporting parsers.
# This handles three common layouts:
#
#   (A) Flat  — all .py files alongside server.py          ← preferred
#   (B) Package — lineage_extraction/{parsers,models}/     ← legacy layout
#
# By adding all candidate dirs to sys.path, flat imports like
# "from reporting_extractor import ..." work in both layouts.
for _subdir in (
    "lineage_extraction",
    "lineage_extraction/parsers",
    "lineage_extraction/models",
    "parsers",
    "models",
):
    _p = PROJECT_ROOT / _subdir
    if _p.is_dir() and str(_p) not in sys.path:
        sys.path.insert(0, str(_p))
del _subdir, _p

from column_lineage_record import ColumnLineageRecord, JobSummaryRecord
from column_lineage_extractor import (
    process_file_sync,
    build_excel_bytes,
    _ops_from_mappings,
)

# ---------------------------------------------------------------------------
# Old extractor imports — used by /extract/report
# ---------------------------------------------------------------------------
# NOTE: Do NOT do a module-level import here.  Any failure in the chain
# (missing pyyaml, un-fixed parser, etc.) used to silently set the flag False
# and the 501 gave no clue.  Now we import lazily inside the endpoint so the
# real exception message surfaces in the HTTP response.
import csv
_REPORT_IMPORT_ERROR: str = ""
_report_process_file = None
ReportingRecord = None


def _ensure_report_imports() -> str:
    global _report_process_file, ReportingRecord, _REPORT_IMPORT_ERROR
    if _report_process_file is not None:
        return ""
    if _REPORT_IMPORT_ERROR:
        return _REPORT_IMPORT_ERROR

    # Belt-and-suspenders: re-add all candidate dirs in case anything was
    # missed at startup (e.g. the subdirs didn't exist yet when server started).
    for _sub in (
        "", "lineage_extraction", "lineage_extraction/parsers",
        "lineage_extraction/models", "parsers", "models",
    ):
        _p = str(PROJECT_ROOT / _sub) if _sub else str(PROJECT_ROOT)
        if _p not in sys.path:
            sys.path.insert(0, _p)

    try:
        from reporting_extractor import _process_file as _rpf
        from reporting_record import ReportingRecord as _RR
        _report_process_file = _rpf
        ReportingRecord = _RR
        return ""
    except Exception as exc:
        import traceback as _tb
        _REPORT_IMPORT_ERROR = (
            f"{type(exc).__name__}: {exc}\n"
            f"sys.path searched: {[p for p in sys.path if 'lineage' in p.lower() or p == str(PROJECT_ROOT)]}\n"
            "Fix: place reporting_extractor.py, reporting_record.py, "
            "looker_parser.py, powerbi_parser.py, qlik_parser.py, "
            "metadata_extractor.py in the same folder as server.py "
            "(or in lineage_extraction/parsers/ / lineage_extraction/models/). "
            "Also run: pip install pyyaml openpyxl"
        )
        return _REPORT_IMPORT_ERROR

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = FastAPI(
    title="Lineage Extraction Tool",
    description=(
        "Column-level data lineage extraction from SQL, PySpark, and Notebook files.\n"
        "Supports file upload and Git repository connection."
    ),
    version="2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

# Serve the frontend
UI_FILE = PROJECT_ROOT / "lineage_extraction_tool.html"
if UI_FILE.exists():
    @app.get("/")
    async def serve_ui():
        return FileResponse(str(UI_FILE))


# ---------------------------------------------------------------------------
# AI Job Description Configuration
# =========================================================================
# Supports multiple AI API providers. Configure ONE of the following:
# 
# 1. GROK (Groq) — Free tier available
#    Provider: "grok"
#    URL: https://api.groq.com/openai/v1/chat/completions
#    Model: llama-3.1-8b-instant
#    Key: Get from https://console.groq.com/keys (starts with: gsk_)
#
# 2. OPENAI (ChatGPT) — Paid, reliable
#    Provider: "openai"
#    URL: https://api.openai.com/v1/chat/completions
#    Model: gpt-4-turbo or gpt-3.5-turbo
#    Key: Get from https://platform.openai.com/api-keys (starts with: sk-)
#
# 3. CLAUDE (Anthropic) — Paid, excellent reasoning
#    Provider: "claude"
#    URL: https://api.anthropic.com/v1/messages
#    Model: claude-3-5-sonnet-20241022 or claude-3-opus-20240229
#    Key: Get from https://console.anthropic.com/keys (starts with: sk-ant-)
#
# 4. CUSTOM OPENAI-COMPATIBLE API — Local or third-party
#    Provider: "openai-compatible"
#    URL: Your custom endpoint URL
#    Model: Your model name
#    Key: Your API key
#
# =========================================================================

# UPDATE THESE SETTINGS FOR YOUR PREFERRED AI PROVIDER
AI_PROVIDER = "grok"  # Options: "grok", "openai", "claude", "openai-compatible"
AI_API_URL = "https://api.groq.com/openai/v1/chat/completions"
AI_MODEL = "llama-3.1-8b-instant"
AI_API_KEY = os.environ.get("AI_API_KEY", "")


async def _ai_job_description(
    file_path: str,
    file_type: str,
    job_name: str,
    source_tables: str,
    target_tables: str,
    operations: str,
    code_snippet: str,
) -> str:
    """Generate AI pipeline description using configured provider.

    Supports: Grok, OpenAI, Claude, and OpenAI-compatible APIs.
    Falls back to static template if key is missing or API call fails.
    """
    if not AI_API_KEY or AI_API_KEY.strip() == "":
        return (
            f"Pipeline '{job_name}' ({file_type}) reads from "
            f"{source_tables or 'unknown source'} and writes to "
            f"{target_tables or 'unknown target'}. "
            f"Operations: {operations or 'none'}. "
            f"(Configure AI_API_KEY in server.py to enable AI descriptions.)"
        )

    prompt = (
        "You are a data engineering documentation assistant.\n\n"
        "Analyse this data pipeline and write a concise 2-3 sentence plain-English "
        "description of what it does. Focus on: what data it reads, what "
        "transformation it applies, and what it produces.\n\n"
        f"File: {file_path}\n"
        f"Type: {file_type}\n"
        f"Job name: {job_name}\n"
        f"Source tables: {source_tables or 'unknown'}\n"
        f"Target tables: {target_tables or 'unknown'}\n"
        f"Operations: {operations or 'none detected'}\n\n"
        f"Code (first 600 chars):\n{code_snippet[:600]}\n\n"
        "Write only the description. No preamble, no bullets, no headers."
    )

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Route to correct provider
            if AI_PROVIDER.lower() == "claude":
                return await _call_claude_api(client, prompt)
            elif AI_PROVIDER.lower() in ("openai", "grok", "openai-compatible"):
                return await _call_openai_compatible_api(client, prompt)
            else:
                print(f"[AI API] Unknown provider: {AI_PROVIDER}", flush=True)
                return _fallback_description(job_name, file_type, source_tables, target_tables, operations)

    except Exception as exc:
        print(f"[AI API] request failed: {exc}", flush=True)
        return _fallback_description(job_name, file_type, source_tables, target_tables, operations)


async def _call_openai_compatible_api(client, prompt: str) -> str:
    """Call OpenAI-compatible APIs: Grok, OpenAI, or custom endpoints."""
    resp = await client.post(
        AI_API_URL,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {AI_API_KEY}",
        },
        json={
            "model": AI_MODEL,
            "max_tokens": 300,
            "messages": [{"role": "user", "content": prompt}],
        },
    )

    if resp.status_code == 200:
        data = resp.json()
        choice = data.get("choices", [{}])[0]
        message = choice.get("message", {})
        content = message.get("content", "").strip()
        if content:
            return content

    print(f"[{AI_PROVIDER.upper()} API] status={resp.status_code} body={resp.text[:200]}", flush=True)
    return None


async def _call_claude_api(client, prompt: str) -> str:
    """Call Claude API (Anthropic)."""
    resp = await client.post(
        AI_API_URL,
        headers={
            "x-api-key": AI_API_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": AI_MODEL,
            "max_tokens": 300,
            "messages": [{"role": "user", "content": prompt}],
        },
    )

    if resp.status_code == 200:
        data = resp.json()
        content = data.get("content", [{}])[0].get("text", "").strip()
        if content:
            return content

    print(f"[CLAUDE API] status={resp.status_code} body={resp.text[:200]}", flush=True)
    return None


def _fallback_description(job_name: str, file_type: str, source_tables: str, target_tables: str, operations: str) -> str:
    """Generate a static description when AI API is unavailable."""
    return (
        f"Pipeline '{job_name}' ({file_type}) reads from "
        f"{source_tables or 'unknown source'} and writes to "
        f"{target_tables or 'unknown target'}. "
        f"Operations detected: {operations or 'none'}."
    )


# ---------------------------------------------------------------------------
# Core processing helper
# ---------------------------------------------------------------------------
async def _process_files_to_excel(
    file_map: dict[str, bytes],
    base_path_prefix: str = "",
    include_ai_descriptions: bool = True,
    include_dependencies: bool = False,
) -> tuple[bytes, int, int, list[str]]:
    """Process a dict of {filename: bytes} and return (xlsx_bytes, n_lineage, n_jobs, errors)."""

    all_lineage: list[ColumnLineageRecord] = []
    all_summary: list[JobSummaryRecord] = []
    errors: list[str] = []

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp = Path(tmp_dir)

        # Write all files to disk so parsers can read them normally
        for filename, content in file_map.items():
            dest = tmp / filename
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(content)

        for filename, content_bytes in file_map.items():
            fp = tmp / filename
            fp_str = (
                f"{base_path_prefix.rstrip('/')}/{filename}"
                if base_path_prefix
                else filename
            )
            job_name = Path(filename).stem

            mappings, file_type, src_tbls, tgt_tbls, error = process_file_sync(fp, fp_str, include_dependencies=include_dependencies)

            if error:
                errors.append(error)
                continue

            _REPORTING_TYPES = {"Qlik","Looker","PowerBI","Tableau","SSRS"}
            for m in mappings:
                tgt_table = m["target_table"]
                tgt_db    = m["target_database"]
                if not tgt_table:
                    if file_type in _REPORTING_TYPES:
                        tgt_table = job_name
                        tgt_db    = "report"
                    else:
                        continue
                all_lineage.append(ColumnLineageRecord(
                    file_path=fp_str,
                    file_type=file_type,
                    job_name=job_name,
                    source_database=m["source_database"],
                    source_table=m["source_table"],
                    source_column=m["source_column"],
                    target_database=tgt_db,
                    target_table=tgt_table,
                    target_column=m["target_column"],
                    sql_operation=m["sql_operation"],
                ))

            # AI description via Grok
            code_snippet = content_bytes.decode("utf-8", errors="replace")[:600]
            if include_ai_descriptions:
                desc = await _ai_job_description(
                    fp_str, file_type, job_name,
                    src_tbls, tgt_tbls,
                    _ops_from_mappings(mappings),
                    code_snippet,
                ) or _fallback_description(job_name, file_type, src_tbls, tgt_tbls, _ops_from_mappings(mappings))
            else:
                desc = (
                    f"Pipeline reads from {src_tbls or 'unknown'} "
                    f"and writes to {tgt_tbls or 'unknown'}. "
                    f"Operations: {_ops_from_mappings(mappings) or 'none'}."
                )

            all_summary.append(JobSummaryRecord(
                file_path=fp_str,
                file_type=file_type,
                job_name=job_name,
                source_tables=src_tbls,
                target_tables=tgt_tbls,
                operations_summary=_ops_from_mappings(mappings),
                job_description=desc,
            ))

    xlsx = build_excel_bytes(all_lineage, all_summary)
    return xlsx, len(all_lineage), len(all_summary), errors, []


# ---------------------------------------------------------------------------
# Endpoint 1: File Upload
# ---------------------------------------------------------------------------
@app.post(
    "/extract/upload",
    summary="Upload pipeline files and extract column-level lineage",
    response_description="Excel file: Sheet 1 = Column Lineage, Sheet 2 = Job Summary",
)
async def extract_upload(
    files: List[UploadFile] = File(
        ..., description=".sql, .py, or .ipynb files"
    ),
    include_ai_descriptions: bool = Form(
        True, description="Generate Grok AI descriptions for Sheet 2"
    ),
    include_dependencies: bool = Form(
        False, description="Also include FILTER_DEPENDENCY and JOIN_KEY rows for full lineage"
    ),
):
    """
    Upload one or more pipeline files (.sql / .py / .ipynb).

    Returns an Excel workbook with:
      - Sheet 1 "Column Lineage": one row per source_column → target_column mapping
      - Sheet 2 "Job Summary": one row per file with Grok AI-generated description
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    file_map: dict[str, bytes] = {}
    for upload in files:
        content = await upload.read()
        file_map[upload.filename] = content

    xlsx, n_lineage, n_jobs, errors, _dup_rows_ignored = await _process_files_to_excel(
        file_map,
        base_path_prefix="",
        include_ai_descriptions=include_ai_descriptions,
        include_dependencies=include_dependencies,
    )

    # Always return the Excel even if n_lineage==0 so the UI can render it.
    # The UI already shows a warning banner when nL==0 — do not block with 422.
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=lineage_output.xlsx",
            "X-Lineage-Rows":  str(n_lineage),
            "X-Summary-Rows":  str(n_jobs),
            "X-Errors-Count":  str(len(errors)),
        },
    )


# ---------------------------------------------------------------------------
# Helper — write Excel workbook to bytes (used by ETL + Report endpoints)
# ---------------------------------------------------------------------------
def _records_to_excel_bytes(sheets: dict, col_headers: list) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(col_headers)
        for row in rows:
            ws.append([row.get(col, "") for col in col_headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _records_to_csv_bytes(rows: list, col_headers: list) -> bytes:
    import csv as _csv
    buf = io.StringIO()
    writer = _csv.DictWriter(buf, fieldnames=col_headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")





# ---------------------------------------------------------------------------
# Endpoint: Reporting lineage extraction → Excel (per BI tool sheets)
# ---------------------------------------------------------------------------
@app.post("/extract/report", summary="Extract BI report lineage from .pbit / .lkml / .qvs")
async def extract_report(
    files: List[UploadFile] = File(..., description=".pbit, .lkml, or .qvs files"),
):
    err = _ensure_report_imports()
    if err:
        raise HTTPException(status_code=501, detail=f"Report extraction unavailable. {err}")
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    all_records = []
    all_warnings: list[str] = []
    all_errors: list[str] = []

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        for upload in files:
            dest = tmp_path / upload.filename
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(await upload.read())

        for file_path in sorted(tmp_path.rglob("*")):
            if not file_path.is_file():
                continue
            try:
                record, warnings, error = _report_process_file(file_path)
                if record:
                    all_records.append(record)
                all_warnings.extend(warnings)
                if error:
                    all_errors.append(error)
            except Exception:
                all_errors.append(f"{file_path.name}: {traceback.format_exc()}")

    if not all_records and all_errors:
        raise HTTPException(status_code=422, detail={"errors": all_errors, "warnings": all_warnings})

    col_headers = ReportingRecord.column_headers()
    sheets: dict = {}
    for rec in all_records:
        sheets.setdefault(rec.tool_name, []).append(rec.to_row())
    if not sheets:
        sheets["Summary"] = [{"domain": "No records extracted"}]

    return StreamingResponse(
        io.BytesIO(_records_to_excel_bytes(sheets, col_headers)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=reporting_lineage_output.xlsx",
            "X-Records-Count": str(len(all_records)),
            "X-Warnings-Count": str(len(all_warnings)),
            "X-Errors-Count": str(len(all_errors)),
        },
    )


# ---------------------------------------------------------------------------
# Endpoint: Git Repository — routes to correct pipeline
# ---------------------------------------------------------------------------
class GitRequest(BaseModel):
    repo_url:    str
    branch:      str       = "main"
    path_filter: str       = ""
    file_types:  list[str] = [".sql", ".py", ".ipynb"]
    include_ai_descriptions: bool = True
    include_dependencies:    bool = False
    pipeline:    str       = "lineage"  # "lineage" | "report"


@app.post(
    "/extract/from-git",
    summary="Connect to a Git repository and extract column-level lineage",
    response_description="Excel file: Sheet 1 = Column Lineage, Sheet 2 = Job Summary",
)
async def extract_from_git(body: GitRequest):
    """
    Connect to a Git repository, fetch matching files, and extract lineage.
    Uses subprocess git clone — no gitpython package required.
    Only requires Git to be installed on the machine (git.exe on Windows).
    """
    import subprocess
    import re
    from fnmatch import fnmatch

    # Check git is available on the system
    try:
        result = subprocess.run(
            ["git", "--version"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            raise HTTPException(
                status_code=501,
                detail="Git is not installed or not in PATH. Install Git from https://git-scm.com"
            )
    except FileNotFoundError:
        raise HTTPException(
            status_code=501,
            detail="Git is not installed or not in PATH. Install Git from https://git-scm.com"
        )

    with tempfile.TemporaryDirectory() as tmp_dir:
        repo_path = Path(tmp_dir) / "repo"

        # Clone the repo using subprocess (shallow, single branch — fast)
        clone_cmd = [
            "git", "clone",
            "--depth", "1",
            "--branch", body.branch,
            "--single-branch",
            body.repo_url,
            str(repo_path)
        ]

        try:
            result = subprocess.run(
                clone_cmd,
                capture_output=True,
                text=True,
                timeout=120
            )
            if result.returncode != 0:
                error_msg = result.stderr.strip() or result.stdout.strip()
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to clone repository: {error_msg}"
                )
        except subprocess.TimeoutExpired:
            raise HTTPException(
                status_code=400,
                detail="Git clone timed out. Check the repository URL and try again."
            )

        # Resolve scan root and file matching based on path_filter
        # Three cases:
        #   1. path_filter is an exact file path  → only that one file
        #   2. path_filter is a folder path        → scan that folder
        #   3. path_filter has * or ? globs        → fnmatch across all files
        #   4. path_filter is empty                → scan entire repo
        scan_root = repo_path
        exact_file: Path | None = None

        if body.path_filter:
            clean_filter = body.path_filter.lstrip("/").replace("\\", "/")
            candidate = repo_path / clean_filter

            if candidate.is_file():
                # Exact file path given (e.g. pipelines/finance/revenue.sql)
                exact_file = candidate
            elif candidate.is_dir():
                # Folder path given (e.g. pipelines/finance/)
                scan_root = candidate
            # else: glob pattern — handled below

        # Collect matching files
        matched: list[Path] = []

        if exact_file is not None:
            matched = [exact_file]
        else:
            for fp in sorted(scan_root.rglob("*")):
                if not fp.is_file():
                    continue
                if fp.suffix.lower() not in body.file_types:
                    continue
                if body.path_filter and any(c in body.path_filter for c in ("*", "?")):
                    rel = str(fp.relative_to(repo_path)).replace("\\", "/")
                    if not fnmatch(rel, body.path_filter.lstrip("/")):
                        continue
                matched.append(fp)

        if not matched:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"No matching files found in {body.repo_url} "
                    f"(branch: {body.branch}, filter: '{body.path_filter}', "
                    f"types: {body.file_types})"
                ),
            )

        # Build file_map: {relative_path: bytes}
        file_map: dict[str, bytes] = {}
        for fp in matched:
            rel = str(fp.relative_to(repo_path)).replace("\\", "/")
            file_map[rel] = fp.read_bytes()

        # Build Git URL prefix — strip embedded token from display URL
        clean_url = body.repo_url.rstrip("/").removesuffix(".git")
        clean_url = re.sub(r"https://[^@]+@", "https://", clean_url)
        git_prefix = f"{clean_url}/blob/{body.branch}"
        repo_name = body.repo_url.rstrip("/").split("/")[-1].removesuffix(".git")

        # ── Route to correct pipeline ────────────────────────────────────────
        pipeline = body.pipeline.lower()

        if pipeline == "report":
            # Report pipeline — uses legacy extractor
            err = _ensure_report_imports()
            if err:
                raise HTTPException(status_code=501, detail=f"Report extraction unavailable. {err}")
            all_records_rpt, all_warnings_rpt, all_errors_rpt = [], [], []
            for fp in matched:
                try:
                    record, warnings, error = _report_process_file(fp)
                    if record: all_records_rpt.append(record)
                    all_warnings_rpt.extend(warnings)
                    if error: all_errors_rpt.append(error)
                except Exception: all_errors_rpt.append(f"{fp.name}: {traceback.format_exc()}")
            sheets_rpt: dict = {}
            for rec in all_records_rpt:
                sheets_rpt.setdefault(rec.tool_name, []).append(rec.to_row())
            if not sheets_rpt: sheets_rpt["Summary"] = [{"domain": "No records extracted"}]
            output_bytes_rpt = _records_to_excel_bytes(sheets_rpt, ReportingRecord.column_headers())
            return StreamingResponse(
                io.BytesIO(output_bytes_rpt),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={repo_name}_reporting.xlsx",
                    "X-Records-Count": str(len(all_records_rpt)),
                    "X-Errors-Count":  str(len(all_errors_rpt)),
                    "X-Files-Scanned": str(len(matched)),
                },
            )

        else:
            # Default: lineage pipeline (column-level)
            xlsx, n_lineage, n_jobs, errors, _dup_rows_ignored = await _process_files_to_excel(
                file_map,
                base_path_prefix=git_prefix,
                include_ai_descriptions=body.include_ai_descriptions,
            )

    if pipeline == "lineage":
        return StreamingResponse(
            io.BytesIO(xlsx),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={repo_name}_lineage.xlsx",
                "X-Lineage-Rows":  str(n_lineage),
                "X-Summary-Rows":  str(n_jobs),
                "X-Errors-Count":  str(len(errors)),
                "X-Files-Scanned": str(len(matched)),
            },
        )


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# AI API status check endpoint
# ---------------------------------------------------------------------------
@app.get("/grok-status", summary="Check AI API status")
async def grok_status():
    """
    Returns whether AI_API_KEY is configured and the API is reachable.
    Also returns provider name and model so the UI pill shows the correct label.
    """
    key = AI_API_KEY.strip() if AI_API_KEY else ""

    # Provider display names for the UI pill
    provider_labels = {
        "grok":              "Claude",
        "openai":            "ChatGPT",
        "claude":            "Claude",
        "openai-compatible": "AI",
    }
    provider_label = provider_labels.get(AI_PROVIDER.lower(), AI_PROVIDER.upper())

    if not key:
        return {
            "connected": False,
            "reason": f"{provider_label} API key not set in server.py",
            "provider": AI_PROVIDER,
            "provider_label": provider_label,
            "model": AI_MODEL,
        }

    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            if AI_PROVIDER.lower() == "claude":
                resp = await client.post(
                    AI_API_URL,
                    headers={
                        "x-api-key": key,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json",
                    },
                    json={"model": AI_MODEL, "max_tokens": 1, "messages": [{"role": "user", "content": "hi"}]},
                )
            else:
                resp = await client.post(
                    AI_API_URL,
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {key}"},
                    json={"model": AI_MODEL, "max_tokens": 1, "messages": [{"role": "user", "content": "hi"}]},
                )

        if resp.status_code in (200, 201):
            return {"connected": True,  "reason": f"{provider_label} API reachable", "provider": AI_PROVIDER, "provider_label": provider_label, "model": AI_MODEL}
        elif resp.status_code == 401:
            return {"connected": False, "reason": "Invalid API key",                  "provider": AI_PROVIDER, "provider_label": provider_label, "model": AI_MODEL}
        else:
            return {"connected": False, "reason": f"{provider_label} returned {resp.status_code}", "provider": AI_PROVIDER, "provider_label": provider_label, "model": AI_MODEL}
    except Exception as exc:
        return {"connected": False, "reason": str(exc), "provider": AI_PROVIDER, "provider_label": provider_label, "model": AI_MODEL}


# ---------------------------------------------------------------------------
# Endpoint: End-to-End Lineage — marry lineage Excel + report Excel
# ---------------------------------------------------------------------------
# Logic (from manager transcript):
#   - Input: two Excel files — one from lineage-extract, one from report-extract
#   - lineage Excel has columns: file_path, job_name, source_table, target_table, etc.
#   - report Excel has columns: file_name, report_name, tables (sources), tool_name, etc.
#   - Join key: lineage target_table = report source table (tables column)
#   - Output: unified end-to-end flow showing:
#       source_table → lineage_job → intermediate_table → report_job → report_name
# ---------------------------------------------------------------------------
@app.post("/extract/e2e", summary="Marry lineage + report Excel to build end-to-end lineage")
async def extract_e2e(
    lineage_file: UploadFile = File(..., description="Excel output from lineage-extract (.xlsx)"),
    report_file:  UploadFile = File(..., description="Excel output from report-extract (.xlsx)"),
):
    """
    Joins lineage-extract output with report-extract output to build
    a complete source → pipeline → table → report flow.

    Join logic:
      lineage target_table  ←matches→  report tables (source tables for the report)

    Output Excel has one sheet 'End-to-End Lineage' with columns:
      source_database, source_table, source_column,
      lineage_job, lineage_file_path, lineage_operation,
      intermediate_table (target of lineage = source of report),
      report_job, report_file, report_name, tool_name
    """
    import openpyxl

    def read_excel_sheet(content: bytes) -> dict[str, list[dict]]:
        """Read all sheets from Excel bytes into {sheet_name: [row_dicts]}."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        result = {}
        for shname in wb.sheetnames:
            ws = wb[shname]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                result[shname] = []
                continue
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
            result[shname] = [
                {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
                for row in rows[1:]
            ]
        return result

    lin_bytes = await lineage_file.read()
    rpt_bytes = await report_file.read()

    try:
        lin_sheets = read_excel_sheet(lin_bytes)
        rpt_sheets = read_excel_sheet(rpt_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel files: {exc}")

    # Get lineage rows — try 'Column Lineage' sheet first, then first sheet
    lin_rows = lin_sheets.get("Column Lineage") or lin_sheets.get(list(lin_sheets.keys())[0]) if lin_sheets else []
    # Get report rows — combine all sheets
    rpt_rows = []
    for rows in rpt_sheets.values():
        rpt_rows.extend(rows)

    if not lin_rows:
        raise HTTPException(status_code=422, detail="No lineage records found in the lineage Excel file.")
    if not rpt_rows:
        raise HTTPException(status_code=422, detail="No report records found in the report Excel file.")

    # Clean lineage rows — fix malformed column names and infer missing source info
    # for AGGREGATE rows where the SQL parser loses the table reference
    def clean_lin_rows(rows: list[dict]) -> list[dict]:
        # Step 1 — strip bracket artifacts from column names
        cleaned = []
        for r in rows:
            col = (r.get("source_column") or "").strip().strip(")(").strip()
            if not col:
                continue
            cleaned.append({**r,
                "source_column":   col,
                "source_table":    r.get("source_table") or "",
                "source_database": r.get("source_database") or "",
            })

        # Step 2 — for AGGREGATE rows with blank source_table,
        # infer from the most common source_table in the same job+target_table group
        # e.g. SUM(t.quantity) comes from retail_transactions (most frequent in that group)
        from collections import Counter

        # Build: (job_name, target_table) → Counter of non-empty source tables
        group_src: dict[tuple, Counter] = {}
        for r in cleaned:
            if r.get("source_table"):
                # lineage Excel uses "job_name" column — not "lineage_job"
                key = (r.get("job_name","") or r.get("lineage_job",""), r.get("target_table",""))
                group_src.setdefault(key, Counter())
                group_src[key][r["source_table"]] += 1
        # Also build db lookup: source_table → source_database
        tbl_db: dict[str, str] = {}
        for r in cleaned:
            if r.get("source_table") and r.get("source_database"):
                tbl_db[r["source_table"]] = r["source_database"]

        # Apply inference
        result = []
        for r in cleaned:
            if not r.get("source_table") and r.get("sql_operation") == "AGGREGATE":
                key = (r.get("job_name","") or r.get("lineage_job",""), r.get("target_table",""))
                counter = group_src.get(key, Counter())
                if counter:
                    best_tbl = counter.most_common(1)[0][0]
                    r = {**r,
                        "source_table":    best_tbl,
                        "source_database": tbl_db.get(best_tbl, ""),
                    }
            result.append(r)

        # Step 3 — deduplicate: same source_table+column+target_table+operation
        seen = set()
        deduped = []
        for r in result:
            key = (r.get("source_table",""), r.get("source_column",""),
                   r.get("target_table",""), r.get("sql_operation",""))
            if key not in seen:
                seen.add(key)
                deduped.append(r)
        return deduped

    lin_rows = clean_lin_rows(lin_rows)

    # Build report lookup: tables → report rows
    # ReportingRecord.tables is a semicolon/comma-separated list of source tables
    def split_tables(val) -> list[str]:
        if not val or str(val).strip().lower() in ("none", ""):
            return []
        return [t.strip().lower() for t in str(val).replace(";", ",").split(",") if t.strip()]

    def bare(name: str) -> str:
        """Strip schema prefix: gold.retail_sales_summary → retail_sales_summary"""
        return name.rsplit(".", 1)[-1].lower() if "." in name else name.lower()

    # Pre-build report table index for fast lookup
    # Index: bare_table_name → [report_rows]
    rpt_index: dict[str, list[dict]] = {}
    for rpt in rpt_rows:
        rpt_tables = split_tables(rpt.get("tables") or rpt.get("TABLES") or "")
        if not rpt_tables:
            # If tables column is empty, try sql_name as fallback
            rpt_tables = split_tables(rpt.get("sql_name") or "")
        for t in rpt_tables:
            key = bare(t)
            rpt_index.setdefault(key, []).append(rpt)

    # e2e join — match lineage target_table to report source tables
    e2e_rows = []
    for lin in lin_rows:
        tgt_raw = (lin.get("target_table") or lin.get("TARGET_TABLE") or "").strip()
        if not tgt_raw:
            continue
        tgt_bare = bare(tgt_raw)

        # Look up matching report rows by bare table name
        matched_rpts = rpt_index.get(tgt_bare, [])

        if matched_rpts:
            for rpt in matched_rpts:
                e2e_rows.append({
                    "source_database":    lin.get("source_database", ""),
                    "source_table":       lin.get("source_table", ""),
                    "source_column":      lin.get("source_column", ""),
                    "lineage_job":        lin.get("job_name", ""),
                    "lineage_file_path":  lin.get("file_path", ""),
                    "lineage_operation":  lin.get("sql_operation", ""),
                    "intermediate_table": tgt_raw,
                    "report_job":         rpt.get("report_name") or rpt.get("file_name", ""),
                    "report_file":        rpt.get("file_name", ""),
                    "report_name":        rpt.get("report_name", ""),
                    "tool_name":          rpt.get("tool_name", ""),
                })
        else:
            # No match — still include row so user can see the lineage flow
            e2e_rows.append({
                "source_database":    lin.get("source_database", ""),
                "source_table":       lin.get("source_table", ""),
                "source_column":      lin.get("source_column", ""),
                "lineage_job":        lin.get("job_name", ""),
                "lineage_file_path":  lin.get("file_path", ""),
                "lineage_operation":  lin.get("sql_operation", ""),
                "intermediate_table": tgt_raw,
                "report_job":         "— no match found",
                "report_file":        "",
                "report_name":        "",
                "tool_name":          "",
            })

    # Count matches
    matched_count = len([r for r in e2e_rows if r["report_job"] != "— no match found"])

    col_headers = [
        "source_database", "source_table", "source_column",
        "lineage_job", "lineage_file_path", "lineage_operation",
        "intermediate_table",
        "report_job", "report_file", "report_name", "tool_name",
    ]

    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb_out = Workbook()
    ws_e2e = wb_out.active
    ws_e2e.title = "End-to-End Lineage"
    ws_e2e.append(col_headers)
    for row in e2e_rows:
        ws_e2e.append([row.get(c, "") for c in col_headers])

    # Also add a summary sheet showing the join stats
    ws_sum = wb_out.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Lineage records (input)",   len(lin_rows)])
    ws_sum.append(["Report records (input)",    len(rpt_rows)])
    ws_sum.append(["E2E matches found",         matched_count])
    ws_sum.append(["Unmatched lineage rows",    len(e2e_rows) - matched_count])
    ws_sum.append(["Total output rows",         len(e2e_rows)])

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=e2e_lineage.xlsx",
            "X-Records-Count":  str(len(e2e_rows)),
            "X-Lineage-Input":  str(len(lin_rows)),
            "X-Report-Input":   str(len(rpt_rows)),
        },
    )




# ---------------------------------------------------------------------------
# Helpers: duplicate detection + circular dependency
# ---------------------------------------------------------------------------
import hashlib
from difflib import SequenceMatcher

def _content_hash(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def _name_similarity(a: str, b: str) -> float:
    """Filename stem similarity (0-1)."""
    return SequenceMatcher(None, Path(a).stem.lower(), Path(b).stem.lower()).ratio()

def _normalise_query(raw: bytes) -> str:
    """Normalise content for duplicate comparison."""
    import re as _re
    text = raw.decode("utf-8", errors="ignore")
    text = _re.sub(r"/\*.*?\*/", " ", text, flags=_re.DOTALL)
    text = _re.sub(r"--[^\n]*", " ", text)
    text = _re.sub(r"//[^\n]*", " ", text)
    text = _re.sub(r"\s+", " ", text).strip().lower()
    return text


def _query_hash(raw: bytes) -> str:
    return hashlib.sha256(_normalise_query(raw).encode()).hexdigest()


def _detect_duplicates(
    file_map: dict[str, bytes]
) -> tuple[dict[str, bytes], list[dict]]:
    """Content-first duplicate detection — filename is irrelevant."""
    seen: dict[str, str] = {}
    dup_rows: list[dict] = []
    keep_map: dict[str, bytes] = {}
    for fname, fbytes in file_map.items():
        h = _query_hash(fbytes)
        if h in seen:
            dup_rows.append({"removed_file": fname, "duplicate_of": seen[h],
                             "reason": "Identical query content (normalised SHA-256 match)",
                             "size_bytes": len(fbytes)})
        else:
            seen[h] = fname
            keep_map[fname] = fbytes
    return keep_map, dup_rows


def _detect_circular(
    lineage_records: list,
) -> list[dict]:
    """
    Find tables that appear as BOTH source_table AND target_table
    within the SAME layer (ETL only, or BI only).
    Tables that are ETL targets AND BI sources are normal E2E flow — not circular.
    """
    _REPORTING_TYPES = {"Qlik", "Looker", "PowerBI", "Tableau", "SSRS"}

    # Sources and targets within ETL layer only
    etl_sources: set[str] = set()
    etl_targets: set[str] = set()
    # Sources within BI layer (i.e. tables the BI tool reads)
    bi_sources: set[str] = set()

    for r in lineage_records:
        ft = getattr(r, "file_type", "") or ""
        st = (getattr(r, "source_table", "") or "").strip().lower()
        tt = (getattr(r, "target_table", "") or "").strip().lower()
        if ft in _REPORTING_TYPES:
            if st:
                bi_sources.add(st)
        else:
            if st:
                etl_sources.add(st)
            if tt:
                etl_targets.add(tt)

    # Only flag tables that are BOTH read AND written within ETL alone
    # (exclude tables that are just ETL→target then BI→source — that is E2E flow)
    circular_tables = etl_sources & etl_targets - bi_sources

    rows: list[dict] = []
    seen: set = set()
    for r in lineage_records:
        ft = getattr(r, "file_type", "") or ""
        if ft in _REPORTING_TYPES:
            continue
        for attr in ("source_table", "target_table"):
            tbl = (getattr(r, attr, "") or "").strip().lower()
            if tbl in circular_tables and tbl not in seen:
                seen.add(tbl)
                rows.append({
                    "table_name": tbl,
                    "appears_as": "source AND target",
                    "warning":    "Circular dependency — this table is read from AND written to within the ETL layer",
                })
    return rows


def _build_table_lineage(lineage_records: list) -> list[dict]:
    """
    Roll up column-level records to table-level:
    one row per unique (source_table → target_table) pair.
    Counts distinct columns, lists operations.

    Dependency rows (target_column=="\u2014") ARE included so that JOIN_KEY tables
    (e.g. accounts used only as a JOIN intermediary) appear in the Table Lineage.
    column_count for dependency rows counts as 0 (they don't add output columns).
    """
    from collections import defaultdict, Counter
    groups: dict[tuple, dict] = defaultdict(lambda: {
        "source_database": "", "source_table": "",
        "target_database": "", "target_table": "",
        "file_types": set(), "job_names": set(),
        "operations": Counter(), "column_count": 0,
    })
    for r in lineage_records:
        key = (
            getattr(r, 'source_table', '') or '',
            getattr(r, 'target_table', '') or '',
        )
        g = groups[key]
        g["source_database"] = getattr(r, 'source_database', '') or g["source_database"]
        g["source_table"]    = key[0]
        g["target_database"] = getattr(r, 'target_database', '') or g["target_database"]
        g["target_table"]    = key[1]
        ft = getattr(r, 'file_type', '') or ''
        jn = getattr(r, 'job_name',  '') or ''
        op = getattr(r, 'sql_operation', '') or ''
        tgt_col = getattr(r, 'target_column', '') or ''
        is_dep  = tgt_col == '\u2014'    # dependency row — no direct output column
        if ft: g["file_types"].add(ft)
        if jn: g["job_names"].add(jn)
        if op: g["operations"][op] += 1
        if not is_dep:
            g["column_count"] += 1        # only count real output column mappings

    rows = []
    for (src_tbl, tgt_tbl), g in groups.items():
        if not src_tbl and not tgt_tbl:
            continue
        rows.append({
            "source_database":  g["source_database"],
            "source_table":     g["source_table"],
            "target_database":  g["target_database"],
            "target_table":     g["target_table"],
            "column_count":     g["column_count"],
            "operations":       ", ".join(sorted(g["operations"].keys())),
            "job_names":        "; ".join(sorted(g["job_names"])),
            "file_types":       "; ".join(sorted(g["file_types"])),
        })
    # Sort: target_table then source_table
    return sorted(rows, key=lambda r: (r["target_table"], r["source_table"]))


# ---------------------------------------------------------------------------
# Endpoint: Unified mixed-file extraction
# Accepts ANY combination of ETL + BI files in one upload.
# Auto-detects format via universal_format_detector.
# Returns ONE Excel with: Table Lineage, Column Lineage, Job Summary,
#                         Circular Dependencies, Duplicates.
# ---------------------------------------------------------------------------
@app.post(
    "/extract/unified",
    summary="Upload any mix of ETL + BI files — auto-detect format, single Excel output",
)
async def extract_unified(
    files: List[UploadFile] = File(..., description="Any pipeline or BI report files — mixed OK"),
    include_ai_descriptions: bool = Form(True),
    include_dependencies: bool = Form(False, description="Include FILTER_DEPENDENCY and JOIN_KEY rows for full lineage"),
):
    """
    Drop any mix of .sql, .py, .ipynb, .pbit, .lkml, .qvs, .yaml, .json, .xml,
    .scala, .r, .sh, .tf etc. The server auto-routes each file through
    universal_format_detector. Duplicate SQLs are detected and logged.
    Circular dependencies are flagged. One Excel is returned with 5 sheets.
    """
    try:
        if not files:
            raise HTTPException(status_code=400, detail="No files uploaded")

        # Read all uploaded files
        raw_map: dict[str, bytes] = {}
        for upload in files:
            raw_map[upload.filename] = await upload.read()

        # ── Step 1: Duplicate detection ──────────────────────────────────────
        clean_map, dup_rows = _detect_duplicates(raw_map)

        # ── Step 2: Column-level lineage extraction (all file types) ─────────
        xlsx_bytes, n_lineage, n_jobs, errors, _dup_rows_ignored = await _process_files_to_excel(
            clean_map,
            base_path_prefix="",
            include_ai_descriptions=include_ai_descriptions,
            include_dependencies=include_dependencies,
        )

        # ── Step 3: Re-parse lineage records for table-level rollup ──────────
        # We need the ColumnLineageRecord objects directly, so re-run sync
        all_lineage_recs: list[ColumnLineageRecord] = []
        all_summary_recs: list[JobSummaryRecord] = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir)
            for fname, content in clean_map.items():
                (tmp / fname).parent.mkdir(parents=True, exist_ok=True)
                (tmp / fname).write_bytes(content)
            for fname, content in clean_map.items():
                fp = tmp / fname
                mappings, file_type, src_tbls, tgt_tbls, error = process_file_sync(fp, fname, include_dependencies=include_dependencies)
                if error:
                    continue
                job_name = Path(fname).stem
                _REPORTING_TYPES = {"Qlik","Looker","PowerBI","Tableau","SSRS"}
                for m in mappings:
                    tgt_table = m["target_table"]
                    tgt_db    = m["target_database"]
                    if not tgt_table:
                        if file_type in _REPORTING_TYPES:
                            tgt_table = job_name
                            tgt_db    = "report"
                        else:
                            continue
                    all_lineage_recs.append(ColumnLineageRecord(
                        file_path=fname, file_type=file_type, job_name=job_name,
                        source_database=m["source_database"],
                        source_table=m["source_table"],
                        source_column=m["source_column"],
                        target_database=tgt_db,
                        target_table=tgt_table,
                        target_column=m["target_column"],
                        sql_operation=m["sql_operation"],
                    ))

        # ── Step 4: Table-level lineage rollup ───────────────────────────────
        table_rows = _build_table_lineage(all_lineage_recs)

        # ── Step 5: Circular dependency detection ────────────────────────────
        circular_rows = _detect_circular(all_lineage_recs)

        # ── Step 6: Build the unified Excel ──────────────────────────────────
        if not OPENPYXL_AVAILABLE:
            raise HTTPException(status_code=500, detail="openpyxl not installed")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        wb.remove(wb.active)

        def _hdr(ws, fill_hex, cols):
            ws.append(cols)
            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = fill; cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sheet 1 — Table Lineage (summary, drill-up view)
        ws_tbl = wb.create_sheet("Table Lineage")
        tbl_hdrs = ["source_database","source_table","target_database","target_table",
                    "column_count","operations","job_names","file_types"]
        _hdr(ws_tbl, "185FA5", tbl_hdrs)
        for r in table_rows:
            ws_tbl.append([r.get(h,"") for h in tbl_hdrs])
        ws_tbl.freeze_panes = "A2"

        # Sheet 2 — Column Lineage (detail view)
        ws_col = wb.create_sheet("Column Lineage")
        col_hdrs = ColumnLineageRecord.column_headers()
        _hdr(ws_col, "0C447C", col_hdrs)
        for r in all_lineage_recs:
            ws_col.append([r.to_row().get(h,"") for h in col_hdrs])
        ws_col.freeze_panes = "A2"

        # Sheet 3 — Job Summary
        ws_sum = wb.create_sheet("Job Summary")
        sum_hdrs = JobSummaryRecord.column_headers()
        _hdr(ws_sum, "2E7D32", sum_hdrs)
        # Re-use the summary records generated in _process_files_to_excel
        # by re-parsing the xlsx_bytes we already built
        try:
            import openpyxl as _opx
            _wb2 = _opx.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
            if "Job Summary" in _wb2.sheetnames:
                _ws2 = _wb2["Job Summary"]
                _rows2 = list(_ws2.iter_rows(values_only=True))
                if len(_rows2) > 1:
                    for _row in _rows2[1:]:
                        ws_sum.append([str(v) if v is not None else "" for v in _row])
        except Exception:
            pass
        ws_sum.freeze_panes = "A2"

        # Sheet 4 — Circular Dependencies
        ws_circ = wb.create_sheet("Circular Dependencies")
        circ_hdrs = ["table_name","appears_as","warning"]
        _hdr(ws_circ, "B71C1C", circ_hdrs)
        for r in circular_rows:
            ws_circ.append([str(r.get(h,"")) for h in circ_hdrs])
        # No placeholder — empty sheet shows 0 in sheet tab
        # (was: ws_circ.append(["—","—","No circular deps"]))
        ws_circ.freeze_panes = "A2"

        # Sheet 5 — Duplicates
        ws_dup = wb.create_sheet("Duplicates")
        dup_hdrs = ["removed_file","duplicate_of","reason","size_bytes"]
        _hdr(ws_dup, "E65100", dup_hdrs)
        for r in dup_rows:
            ws_dup.append([str(r.get(h,"")) for h in dup_hdrs])
        # No placeholder — empty sheet shows 0 in sheet tab
        # (was: ws_dup.append(["—","—","No duplicates"]))
        ws_dup.freeze_panes = "A2"

        # Sheet 6 — E2E Lineage
        # Correct join: ETL.target_table == PBIT.source_table
        #           AND ETL.target_column == PBIT.source_column
        # This produces one row per source→intermediate_col→report_col mapping.
        # A cross-join (all ETL rows × all PBIT cols) is WRONG and produces
        # spurious rows like "transaction_date flows into region".
        _REPORTING_TYPES_E2E = {"Qlik", "Looker", "PowerBI", "Tableau", "SSRS"}
        etl_rows_e2e = [r for r in all_lineage_recs if r.file_type not in _REPORTING_TYPES_E2E]
        rpt_rows_e2e = [r for r in all_lineage_recs if r.file_type in _REPORTING_TYPES_E2E]
        e2e_rows: list[dict] = []
        if etl_rows_e2e and rpt_rows_e2e:
            # Index PBIT rows by (source_table_lower, source_column_lower)
            rpt_idx: dict[tuple, list] = {}
            for rr in rpt_rows_e2e:
                key = (
                    (rr.source_table  or "").strip().lower(),
                    (rr.source_column or "").strip().lower(),
                )
                if key[0]:
                    rpt_idx.setdefault(key, []).append(rr)

            seen_e2e: set[tuple] = set()
            for er in etl_rows_e2e:
                # Match on ETL target_table + target_column == PBIT source_table + source_column
                key = (
                    (er.target_table  or "").strip().lower(),
                    (er.target_column or "").strip().lower(),
                )
                for rr in rpt_idx.get(key, []):
                    dedup_key = (
                        er.source_table, er.source_column,
                        er.target_table, er.target_column,
                        rr.job_name, rr.source_column,
                    )
                    if dedup_key not in seen_e2e:
                        seen_e2e.add(dedup_key)
                        e2e_rows.append({
                            "source_database":    er.source_database,
                            "source_table":       er.source_table,
                            "source_column":      er.source_column,
                            "etl_job":            er.job_name,
                            "etl_file":           er.file_path,
                            "etl_file_type":      er.file_type,
                            "etl_operation":      er.sql_operation,
                            "intermediate_table": er.target_table,
                            "intermediate_column": er.target_column,
                            "report_job":         rr.job_name,
                            "report_file":        rr.file_path,
                            "report_file_type":   rr.file_type,
                            "report_column":      rr.source_column,
                        })

        ws_e2e = wb.create_sheet("E2E Lineage")
        e2e_hdrs = [
            "source_database", "source_table", "source_column",
            "etl_job", "etl_file", "etl_file_type", "etl_operation",
            "intermediate_table", "intermediate_column",
            "report_job", "report_file", "report_file_type", "report_column",
        ]
        _hdr(ws_e2e, "4A148C", e2e_hdrs)
        for row in e2e_rows:
            ws_e2e.append([str(row.get(h, "")) for h in e2e_hdrs])
        ws_e2e.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        n_tables = len(table_rows)
        n_circular = len(circular_rows)
        n_dups = len(dup_rows)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":  "attachment; filename=unified_lineage.xlsx",
                "X-Lineage-Rows":       str(n_lineage),
                "X-Table-Rows":         str(n_tables),
                "X-Summary-Rows":       str(n_jobs),
                "X-Circular-Count":     str(n_circular),
                "X-Duplicate-Count":    str(n_dups),
                "X-E2E-Count":          str(len(e2e_rows)),
                "X-Errors-Count":       str(len(errors)),
            },
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"{type(exc).__name__}: {exc} | {traceback.format_exc()}",
        )

# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------
@app.get("/health", summary="Health check")
async def health():
    return {"status": "ok", "version": "2.0.0"}


# ---------------------------------------------------------------------------
# Dev entrypoint
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=True)
