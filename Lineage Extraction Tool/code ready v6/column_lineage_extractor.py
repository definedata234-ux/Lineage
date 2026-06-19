"""Column-level lineage extraction orchestrator.

Replaces lineage_extractor.py and etl_extractor.py with a single unified
pipeline that produces:

  Sheet 1 "Column Lineage"  — one row per source_column → target_column mapping
  Sheet 2 "Job Summary"     — one row per file, AI-generated description

Per manager feedback:
  - domain removed → file_path is the primary identifier
  - No UDF / delta / complexity columns
  - Column-level granularity (not table-level)
  - AI job description in a separate sheet
  - File upload AND Git repository as input modes (Git via server endpoint)

File types handled:
  .sql   → column_sql_parser.extract_column_mappings()
  .py    → column_pyspark_parser.extract_pyspark_column_mappings()
  .ipynb → code cells concatenated → same as .py

Error handling: per-file errors are caught and logged; processing continues.
"""

import ast
import io
import json
import re
import sys
import traceback
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from column_lineage_record import ColumnLineageRecord, JobSummaryRecord
from column_sql_parser import extract_column_mappings, ColumnMapping
from column_pyspark_parser import extract_pyspark_column_mappings
from universal_format_detector import extract_all_content, get_file_type_label


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Header style colours for Excel
_HEADER_FILL_LINEAGE = "185FA5"   # blue — Sheet 1
_HEADER_FILL_SUMMARY = "2E7D32"  # green — Sheet 2
_HEADER_FONT_COLOR   = "FFFFFF"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ops_from_mappings(mappings: list[ColumnMapping]) -> str:
    """Return comma-separated unique sql_operation values."""
    ops = list(dict.fromkeys(m["sql_operation"] for m in mappings))
    return ", ".join(ops)


def _all_tables(mappings: list[ColumnMapping], key: str) -> str:
    """Semicolon-separated unique values of mappings[key]."""
    vals = list(dict.fromkeys(m[key] for m in mappings if m[key]))
    return ";".join(vals)


# ---------------------------------------------------------------------------
# Core file processor  (universal — all file formats)
# ---------------------------------------------------------------------------

def process_file_sync(
    file_path: Path,
    file_path_str: str,
    include_dependencies: bool = False,
) -> tuple[list[ColumnMapping], str, str, str, Optional[str]]:
    """Process ANY file format and return column-level lineage mappings.

    Uses universal_format_detector to identify the format and extract
    SQL/code chunks, then routes each chunk to the appropriate column parser.

    Args:
        file_path:     Path on disk (used for reading and format detection).
        file_path_str: String stored in the output — local path or Git URL.

    Returns:
        (mappings, file_type_label, source_tables_str, target_tables_str, error)
        error is None on success.
    """
    try:
        content_bytes = file_path.read_bytes()
    except OSError as exc:
        return ([], "Unknown", "", "", f"Failed to read {file_path.name}: {exc}")

    # Detect format and extract all content chunks
    try:
        chunks = extract_all_content(file_path, content_bytes)
        file_type = get_file_type_label(file_path, content_bytes)
    except Exception as exc:
        return ([], "Unknown", "", "", f"Format detection failed for {file_path.name}: {exc}")

    if not chunks:
        return ([], file_type, "", "", f"No extractable content in {file_path.name}")

    mappings: list[ColumnMapping] = []

    for chunk in chunks:
        dialect = chunk["dialect"]
        content  = chunk["content"]
        if not content.strip():
            continue

        try:
            if dialect == "sql":
                mappings.extend(extract_column_mappings(content, include_dependencies=include_dependencies))

            elif dialect == "python":
                mappings.extend(extract_pyspark_column_mappings(content, file_path_str))
                try:
                    tree = ast.parse(content)
                    for node in ast.walk(tree):
                        if not isinstance(node, ast.Call):
                            continue
                        f = node.func
                        if (isinstance(f, ast.Attribute) and f.attr == "sql"
                                and isinstance(f.value, ast.Name) and f.value.id == "spark"):
                            if node.args and isinstance(node.args[0], ast.Constant):
                                sql_str = node.args[0].value
                                if isinstance(sql_str, str):
                                    mappings.extend(extract_column_mappings(sql_str, include_dependencies=include_dependencies))
                except SyntaxError:
                    pass

            elif dialect in ("scala", "java"):
                for m in re.finditer(r'spark\.sql\s*\(\s*"""(.*?)"""\s*\)', content, re.DOTALL):
                    mappings.extend(extract_column_mappings(m.group(1), include_dependencies=include_dependencies))
                for m in re.finditer(r'spark\.sql\s*\(\s*"((?:[^"\\]|\\.)*)"\s*\)', content):
                    mappings.extend(extract_column_mappings(m.group(1), include_dependencies=include_dependencies))

            elif dialect == "r":
                if re.search(r"\b(SELECT|INSERT|CREATE)\b", content, re.IGNORECASE):
                    mappings.extend(extract_column_mappings(content, include_dependencies=include_dependencies))

            elif dialect == "lookml":
                if re.search(r"\b(SELECT|FROM)\b", content, re.IGNORECASE):
                    mappings.extend(extract_column_mappings(content, include_dependencies=include_dependencies))

            elif dialect == "qlik":
                if re.search(r"\b(SELECT|FROM)\b", content, re.IGNORECASE):
                    mappings.extend(extract_column_mappings(content, include_dependencies=include_dependencies))

            elif dialect in ("powerbi", "ssrs"):
                # PowerBI .pbit / SSRS .rdl — no embedded SQL in many files.
                # Extract column-level lineage directly from the dataModel:
                #   source_table = table name (e.g. retail_sales_summary)
                #   source_column = column name
                #   target_table = report file stem (job_name)
                #   target_database = "report"
                # This ensures E2E join works even when no SQL queries are embedded.
                try:
                    import json as _json
                    _data = _json.loads(content)
                    _report_name = file_path.stem
                    _dm = _data.get("dataModel", _data.get("model", {}))
                    for _tbl in _dm.get("tables", []):
                        _tbl_name = _tbl.get("name", "").strip()
                        _tbl_src  = _tbl.get("source", _tbl_name)
                        # strip db prefix from source (e.g. gold.retail_sales_summary → retail_sales_summary)
                        _bare_src = _tbl_src.rsplit(".", 1)[-1] if "." in _tbl_src else _tbl_src
                        _src_db   = _tbl_src.rsplit(".", 1)[0]  if "." in _tbl_src else ""
                        for _col in _tbl.get("columns", []):
                            _col_name = _col.get("name", "").strip()
                            if _col_name:
                                mappings.append(ColumnMapping(
                                    source_database=_src_db,
                                    source_table=_bare_src or _tbl_name,
                                    source_column=_col_name,
                                    target_database="report",
                                    target_table=_report_name,
                                    target_column=_col_name,
                                    sql_operation="SELECT",
                                ))
                except Exception:
                    # Fallback: try SQL extraction if JSON parsing fails
                    if re.search(r"\b(SELECT|FROM)\b", content, re.IGNORECASE):
                        mappings.extend(extract_column_mappings(content, include_dependencies=include_dependencies))

            elif dialect in ("yaml", "json", "xml", "hcl", "shell", "unknown"):
                if re.search(r"\b(SELECT|INSERT|CREATE|MERGE)\b", content, re.IGNORECASE):
                    mappings.extend(extract_column_mappings(content, include_dependencies=include_dependencies))

            # binary, data, ini, toml — no column extraction possible; skip

        except Exception as exc:
            # Per-chunk error — log and continue with remaining chunks
            print(f"[extractor] chunk error in {file_path.name} ({dialect}): {exc}", flush=True)
            continue

    src_tables = _all_tables(mappings, "source_table")
    tgt_tables = _all_tables(mappings, "target_table")
    return mappings, file_type, src_tables, tgt_tables, None


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _style_header_row(ws, fill_hex: str) -> None:
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(bold=True, color=_HEADER_FONT_COLOR)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)



import re as _re_cell

def _safe_cell(value) -> str:
    """Strip illegal control characters before writing to an Excel cell.

    openpyxl raises IllegalCharacterError for codepoints 0x00-0x08,
    0x0B-0x0C, 0x0E-0x1F (and 0xFFFE/0xFFFF).  These appear in
    AI-generated descriptions when the LLM API returns garbled bytes.
    """
    if not isinstance(value, str):
        return value if value is not None else ''
    return _re_cell.sub(r"[\u0000-\u0008\u000b\u000c\u000e-\u001f\u007f\ufffe\uffff]", "", value)

def build_excel_bytes(
    lineage_records: list[ColumnLineageRecord],
    summary_records: list[JobSummaryRecord],
) -> bytes:
    """Write both sheets to an Excel workbook and return bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Column Lineage ───────────────────────────────────────────
    ws1 = wb.create_sheet("Column Lineage")
    headers1 = ColumnLineageRecord.column_headers()
    ws1.append(headers1)
    _style_header_row(ws1, _HEADER_FILL_LINEAGE)

    for rec in lineage_records:
        ws1.append([_safe_cell(rec.to_row().get(h, "")) for h in headers1])

    # Column widths
    col_widths = {
        "file_path": 45, "file_type": 10, "job_name": 25,
        "source_database": 18, "source_table": 22, "source_column": 20,
        "target_database": 18, "target_table": 22, "target_column": 20,
        "sql_operation": 14,
    }
    for i, h in enumerate(headers1, 1):
        ws1.column_dimensions[ws1.cell(1, i).column_letter].width = col_widths.get(h, 18)

    # Freeze header row
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Job Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Job Summary")
    headers2 = JobSummaryRecord.column_headers()
    ws2.append(headers2)
    _style_header_row(ws2, _HEADER_FILL_SUMMARY)

    for rec in summary_records:
        ws2.append([_safe_cell(rec.to_row().get(h, "")) for h in headers2])

    sum_widths = {
        "file_path": 45, "file_type": 10, "job_name": 25,
        "source_tables": 30, "target_tables": 30,
        "operations_summary": 25, "job_description": 60,
    }
    for i, h in enumerate(headers2, 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = sum_widths.get(h, 20)
        # Wrap job_description column
        if h == "job_description":
            for row_idx in range(2, ws2.max_row + 1):
                ws2.cell(row_idx, i).alignment = Alignment(wrap_text=True, vertical="top")

    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# CLI entry point (synchronous, no AI descriptions)
# ---------------------------------------------------------------------------

def run_extraction_cli(
    input_dir: str,
    output_path: str,
    patterns: list[str] | None = None,
    git_base_url: str = "",
) -> int:
    """Run column-level lineage extraction from a local directory.

    Args:
        input_dir:    Directory to scan recursively.
        output_path:  Path for the output .xlsx file.
        patterns:     Glob patterns (default: *.sql, *.py, *.ipynb).
        git_base_url: If set, file_path values use this as URL prefix
                      instead of local paths (e.g. for Git-sourced files).

    Returns:
        0 on success, 1 if any file had errors.
    """
    from fnmatch import fnmatch

    patterns = patterns or ["*.sql", "*.py", "*.ipynb"]
    input_path = Path(input_dir)
    files = sorted(
        f for f in input_path.rglob("*")
        if f.is_file() and any(fnmatch(f.name, p) for p in patterns)
    )

    all_lineage: list[ColumnLineageRecord] = []
    all_summary: list[JobSummaryRecord] = []
    errors: list[str] = []

    for fp in files:
        # Compute the file path string for the output
        if git_base_url:
            rel = str(fp.relative_to(input_path)).replace("\\", "/")
            fp_str = f"{git_base_url.rstrip('/')}/{rel}"
        else:
            fp_str = str(fp.relative_to(input_path)).replace("\\", "/")

        job_name = fp.stem
        mappings, file_type, src_tbls, tgt_tbls, error = process_file_sync(fp, fp_str)

        if error:
            print(f"ERROR: {error}", file=sys.stderr)
            errors.append(error)
            continue

        # Build ColumnLineageRecord rows
        for m in mappings:
            all_lineage.append(ColumnLineageRecord(
                file_path=fp_str,
                file_type=file_type,
                job_name=job_name,
                source_database=m["source_database"],
                source_table=m["source_table"],
                source_column=m["source_column"],
                target_database=m["target_database"],
                target_table=m["target_table"],
                target_column=m["target_column"],
                sql_operation=m["sql_operation"],
            ))

        # Build JobSummaryRecord (no AI description in CLI mode)
        all_summary.append(JobSummaryRecord(
            file_path=fp_str,
            file_type=file_type,
            job_name=job_name,
            source_tables=src_tbls,
            target_tables=tgt_tbls,
            operations_summary=_ops_from_mappings(mappings),
            job_description="(Run via server for AI-generated description)",
        ))

    xlsx_bytes = build_excel_bytes(all_lineage, all_summary)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(xlsx_bytes)

    print(f"\n{'='*55}")
    print(f"COLUMN LINEAGE EXTRACTION COMPLETE")
    print(f"{'='*55}")
    print(f"Files processed:     {len(files)}")
    print(f"Column mappings:     {len(all_lineage)}")
    print(f"Job summaries:       {len(all_summary)}")
    print(f"Errors:              {len(errors)}")
    print(f"Output:              {output_path}")
    print(f"{'='*55}")

    return 1 if errors else 0


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Usage: python column_lineage_extractor.py <input_dir> <output.xlsx>")
        sys.exit(1)
    sys.exit(run_extraction_cli(sys.argv[1], sys.argv[2]))
